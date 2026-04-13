from __future__ import annotations

import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = Path('/Users/ava/Documents/GitHub/first-access-lending')
WORKBOOK = SOURCE_ROOT / 'getaccess' / 'ratesheets' / '2026-04-13 - fully delegated - external rate sheet - Intraday.xlsx'
GENERATED = ROOT / 'src' / 'lib' / 'rates' / 'button-ratesheet.json'


def expect(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def main() -> None:
    wb = load_workbook(WORKBOOK, data_only=False)
    ws = wb['Fully Delegated']

    with GENERATED.open('r', encoding='utf-8') as f:
        data = json.load(f)

    expect(len(data['noteRates']) == 40, f"expected 40 note rates, got {len(data['noteRates'])}")
    expect(data['noteRates'][0]['noteRate'] == 12, 'first note rate should be 12')
    expect(data['noteRates'][-1]['noteRate'] == 7.125, 'last note rate should be 7.125')

    expect(data['noteRates'][0]['prices']['fullDoc']['HELOC'] == 108.375, 'first HELOC price mismatch')
    expect(data['noteRates'][0]['prices']['fullDoc']['CES'] == 108.8125, 'first CES price mismatch')

    cltv_rows = data['tables']['cltv']['rows']
    expect(len(cltv_rows) == 9, f'expected 9 cltv rows, got {len(cltv_rows)}')
    expect(cltv_rows[0] == 'FICO 620 - 639', 'unexpected first CLTV row label')

    dti_rows = data['tables']['dti']['rows']
    expect(len(dti_rows) == 7, f'expected 7 dti rows, got {len(dti_rows)}')
    expect(dti_rows[0] == 'HELOC', f'unexpected first DTI row label: {dti_rows[0]}')

    cash_out_rows = data['tables']['cashOut']['rows']
    expect(cash_out_rows[0] == 'Cash Out', 'cash-out row label mismatch')

    occupancy_rows = data['tables']['occupancy']['rows']
    expect(occupancy_rows == ['Second Home', '2-4 Unit', 'Investor'], f'unexpected occupancy rows: {occupancy_rows}')

    # Workbook spot checks so we know the source workbook stayed in sync with the generated snapshot.
    expect(ws['A13'].value == 12, 'workbook note-rate top row mismatch')
    expect(ws['B13'].value == 108.375, 'workbook full-doc HELOC price mismatch')
    expect(ws['C13'].value == 108.8125, 'workbook full-doc CES price mismatch')

    print('Button ratesheet verification passed.')


if __name__ == '__main__':
    main()
