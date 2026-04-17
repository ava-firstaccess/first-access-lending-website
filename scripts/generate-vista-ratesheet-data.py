#!/usr/bin/env python3
import json
from pathlib import Path
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = Path('/Users/ava/Documents/GitHub/first-access-lending/getaccess/ratesheets/latest_vista.xlsx')
OUTPUT_PATH = REPO_ROOT / 'src' / 'lib' / 'rates' / 'vista-ratesheet.json'

PROGRAMS = {
    'secondOO': {
        'key': 'PT',
        'inputCode': 'PR',
        'inputName': 'Second OO',
        'pricingRows': (65, 113),
        'basePriceColumn': 6,
        'maxPriceRows': {'default': 41},
        'minPriceRows': {'default': 52},
        'cltv': {
            'headerRow': 65,
            'startRow': 66,
            'endRow': 95,
            'columns': (20, 27),
            'docGroups': {
                'doc01': [(66, 75)],
                'doc02': [(76, 85)],
                'doc03': [(86, 95)],
            },
        },
        'adjustmentRows': {
            'bankStatement': (106, 108),
            'term': (109, 114),
            'amortization': (115, 116),
            'loanAmount': (117, 137),
            'dti': (138, 142),
            'dscr': (143, 143),
            'purpose': (144, 146),
            'occupancy': (147, 148),
            'valuation': (149, 150),
            'state': (151, 154),
            'propertyType': (155, 169),
            'citizenship': (170, 173),
            'employment': (174, 178),
            'firstTimeBuyer': (179, 184),
            'housingHistory': (185, 190),
            'creditEvents': (191, 197),
            'payment': (198, 199),
            'servicing': (200, 201),
            'prepay': (202, 202),
            'lockTerm': (203, 205),
            'lockType': (206, 207),
        },
    },
    'secondNOO': {
        'key': 'IT',
        'inputCode': 'IR',
        'inputName': 'Second NOO',
        'pricingRows': (220, 269),
        'basePriceColumn': 6,
        'maxPriceRows': {
            'No Prepay - Hard': 295,
            '1yr Prepay - Hard': 296,
            '2yr Prepay - Hard': 299,
            '3yr Prepay - Hard': 300,
            '4yr Prepay - Hard': 301,
            '5yr Prepay - Hard': 302,
        },
        'minPriceRows': {'default': 443},
        'cltv': {
            'headerRow': 220,
            'startRow': 221,
            'endRow': 260,
            'columns': (20, 27),
            'docGroups': {
                'doc01': [(221, 230)],
                'doc02': [(231, 240)],
                'doc04': [(241, 250)],
                'doc05': [(251, 260)],
            },
        },
        'adjustmentRows': {
            'bankStatement': (261, 263),
            'term': (264, 269),
            'amortization': (270, 271),
            'loanAmount': (272, 292),
            'dti': (293, 297),
            'dscr': (298, 304),
            'purpose': (305, 307),
            'occupancy': (308, 308),
            'valuation': (309, 310),
            'state': (311, 314),
            'propertyType': (315, 329),
            'citizenship': (330, 333),
            'employment': (334, 338),
            'firstTimeBuyer': (339, 344),
            'housingHistory': (345, 350),
            'creditEvents': (351, 357),
            'payment': (358, 359),
            'servicing': (360, 361),
            'prepay': (362, 382),
            'lockTerm': (383, 385),
            'lockType': (386, 387),
        },
    },
}


def normalize_value(value):
    if value is None or value == '':
        return None
    if isinstance(value, str):
        text = value.strip()
        if text in {'#N/A', 'na', 'n/a'}:
            return None
        return text
    return value


def price(value):
    value = normalize_value(value)
    if value is None:
        return None
    try:
        return round(float(value), 3)
    except (TypeError, ValueError):
        return None


def parse_pricing_table(ws, start_row, end_row, base_col):
    rows = []
    for r in range(start_row, end_row + 1):
        note_rate = normalize_value(ws.cell(r, 4).value)
        base_price = price(ws.cell(r, base_col).value)
        if isinstance(note_rate, (int, float)) and base_price is not None:
            rows.append({
                'noteRate': round(float(note_rate), 3),
                'basePrice': base_price,
            })
    return rows


def parse_bucket_labels(ws, header_row, start_col, end_col):
    return [str(normalize_value(ws.cell(header_row, c).value)) for c in range(start_col, end_col + 1)]


def parse_credit_scores(ws, start_row, end_row):
    scores = []
    for r in range(start_row, end_row + 1):
        label = normalize_value(ws.cell(r, 14).value)
        if label is not None:
            scores.append(str(label))
    return scores


def parse_cltv_groups(ws, spec):
    labels = parse_bucket_labels(ws, spec['headerRow'], spec['columns'][0], spec['columns'][1])
    groups = {}
    for group_name, ranges in spec['docGroups'].items():
        rows = []
        for start_row, end_row in ranges:
            for r in range(start_row, end_row + 1):
                score_label = str(normalize_value(ws.cell(r, 14).value))
                values = [price(ws.cell(r, c).value) for c in range(spec['columns'][0], spec['columns'][1] + 1)]
                rows.append({'creditScore': score_label, 'values': values})
        groups[group_name] = {
            'cltvBuckets': labels,
            'rows': rows,
        }
    return groups


def parse_adjustment_rows(ws, start_row, end_row, *, matrix=False):
    items = []
    for r in range(start_row, end_row + 1):
        label = normalize_value(ws.cell(r, 14).value)
        lookup = normalize_value(ws.cell(r, 19).value)
        value = price(ws.cell(r, 20).value)
        values = [price(ws.cell(r, c).value) for c in range(20, 28)] if matrix else None
        if label is None and lookup is None and value is None and not (matrix and any(v is not None for v in values or [])):
            continue
        item = {
            'label': str(label) if label is not None else '',
            'lookupKey': str(lookup) if lookup is not None else '',
            'value': value,
        }
        if matrix:
            item['values'] = values
        items.append(item)
    return items


def parse_price_map(ws, row_map):
    out = {}
    for key, row in row_map.items():
        value = None
        for col in range(4, 9):
            candidate = price(ws.cell(row, col).value)
            if candidate is not None:
                value = candidate
                break
        out[key] = value
    return out


def main():
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb['Price']

    data = {
        'sourceWorkbook': str(WORKBOOK_PATH),
        'sheet': 'Price',
        'programs': {},
    }

    for name, spec in PROGRAMS.items():
        data['programs'][name] = {
            'inputCode': spec['inputCode'],
            'inputName': spec['inputName'],
            'sections': {
                'pricing30Day': {
                    'rows': list(spec['pricingRows']),
                    'rowsData': parse_pricing_table(ws, spec['pricingRows'][0], spec['pricingRows'][1], spec['basePriceColumn']),
                    'maxPrice': parse_price_map(ws, spec['maxPriceRows']),
                    'minPrice': parse_price_map(ws, spec['minPriceRows']),
                },
                'cltvDoc01': {
                    'rowRange': [spec['cltv']['docGroups']['doc01'][0][0], spec['cltv']['docGroups']['doc01'][-1][1]],
                    **parse_cltv_groups(ws, spec['cltv'])['doc01'],
                },
                **{
                    f'cltv{group_name.capitalize()}': {
                        'rowRange': [ranges[0][0], ranges[-1][1]],
                        **parse_cltv_groups(ws, spec['cltv'])[group_name],
                    }
                    for group_name, ranges in spec['cltv']['docGroups'].items()
                    if group_name != 'doc01'
                },
                'adjustments': {
                    category: {
                        'rows': list(row_range),
                        'items': parse_adjustment_rows(ws, row_range[0], row_range[1], matrix=(category in {'propertyType', 'dti'})),
                    }
                    for category, row_range in spec['adjustmentRows'].items()
                },
            },
        }

    OUTPUT_PATH.write_text(json.dumps(data, indent=2) + '\n')
    print(f'Wrote {OUTPUT_PATH}')


if __name__ == '__main__':
    main()
