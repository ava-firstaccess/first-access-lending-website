#!/usr/bin/env python3
from __future__ import annotations

import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path('/Users/ava/Documents/GitHub/first-access-lending/getaccess/ratesheets/latest_osb.xlsm')
OUTPUT = ROOT / 'src/lib/rates/osb-ratesheet.json'


def normalize(value):
    if value in (None, ''):
        return None
    if isinstance(value, str):
        value = value.strip()
        return None if value == '-' else value
    return value


def price(value):
    value = normalize(value)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 3)
    return value


wb = load_workbook(WORKBOOK, data_only=True)

second = wb['2nd Liens']
heloc = wb['HELOC']

data = {
    'sourceWorkbook': str(WORKBOOK),
    'programs': {
        'secondLiens': {
            'sheet': '2nd Liens',
            'pricing': {
                'rateType': 'fixed-note-rate',
                'products': [
                    {'key': 'fixed10', 'label': 'Fixed 10'},
                    {'key': 'fixed15', 'label': 'Fixed 15'},
                    {'key': 'fixed20', 'label': 'Fixed 20'},
                    {'key': 'fixed30', 'label': 'Fixed 30'},
                ],
                'rowsData': [
                    {
                        'rate': round(float(second.cell(row, 3).value), 3),
                        'prices': {
                            'fixed10': price(second.cell(row, 4).value),
                            'fixed15': price(second.cell(row, 5).value),
                            'fixed20': price(second.cell(row, 6).value),
                            'fixed30': price(second.cell(row, 7).value),
                        },
                    }
                    for row in range(10, 75)
                    if isinstance(second.cell(row, 3).value, (int, float))
                ],
            },
            'constraints': {
                'minPrice': price(second.cell(77, 5).value),
                'maxPrice30Year': price(second.cell(78, 5).value),
                'maxPriceShorterTerm': price(second.cell(79, 5).value),
            },
            'cltvBuckets': [str(second.cell(9, col).value) for col in range(11, 20)],
            'creditMatrix': [
                {
                    'creditScore': str(second.cell(row, 10).value),
                    'values': [price(second.cell(row, col).value) for col in range(11, 20)],
                }
                for row in range(10, 19)
            ],
            'adjustments': {
                'loanAmount': [
                    {'label': str(second.cell(row, 10).value), 'values': [price(second.cell(row, col).value) for col in range(11, 20)]}
                    for row in range(21, 26)
                ],
                'loanType': [
                    {'label': str(second.cell(row, 10).value), 'values': [price(second.cell(row, col).value) for col in range(11, 20)]}
                    for row in range(26, 34)
                ],
                'property': [
                    {'label': str(second.cell(row, 10).value), 'values': [price(second.cell(row, col).value) for col in range(11, 20)]}
                    for row in range(35, 40)
                ],
            },
            'lockAdjustments': [
                {'label': str(second.cell(row, 24).value), 'value': price(second.cell(row, 25).value)}
                for row in range(19, 22)
            ],
        },
        'heloc': {
            'sheet': 'HELOC',
            'pricing': {
                'rateType': 'margin',
                'products': [
                    {'key': 'heloc20', 'label': '20 Year Maturity'},
                    {'key': 'heloc30', 'label': '30 Year Maturity'},
                ],
                'rowsData': [
                    {
                        'rate': round(float(heloc.cell(row, 3).value), 3),
                        'prices': {
                            'heloc20': price(heloc.cell(row, 4).value),
                            'heloc30': price(heloc.cell(row, 5).value),
                        },
                    }
                    for row in range(10, 65)
                    if isinstance(heloc.cell(row, 3).value, (int, float))
                ],
            },
            'constraints': {
                'minPrice': price(heloc.cell(29, 21).value),
                'maxPrice30Year': price(heloc.cell(30, 21).value),
                'maxPriceShorterTerm': price(heloc.cell(31, 21).value),
            },
            'cltvBuckets': [str(heloc.cell(9, col).value) for col in range(9, 18)],
            'creditMatrix': [
                {
                    'creditScore': str(heloc.cell(row, 8).value),
                    'values': [price(heloc.cell(row, col).value) for col in range(9, 18)],
                }
                for row in range(10, 19)
            ],
            'adjustments': {
                'drawTerm': [
                    {'label': str(heloc.cell(row, 8).value), 'values': [price(heloc.cell(row, col).value) for col in range(9, 18)]}
                    for row in range(21, 24)
                ],
                'loanAmount': [
                    {'label': str(heloc.cell(row, 8).value), 'values': [price(heloc.cell(row, col).value) for col in range(9, 18)]}
                    for row in range(24, 29)
                ],
                'loanType': [
                    {'label': str(heloc.cell(row, 8).value), 'values': [price(heloc.cell(row, col).value) for col in range(9, 18)]}
                    for row in range(29, 33)
                ],
                'property': [
                    {'label': str(heloc.cell(row, 8).value), 'values': [price(heloc.cell(row, col).value) for col in range(9, 18)]}
                    for row in range(33, 40)
                ],
            },
            'lockAdjustments': [
                {'label': str(heloc.cell(row, 23).value), 'value': price(heloc.cell(row, 24).value)}
                for row in range(21, 24)
            ],
            'armFeatures': {
                str(heloc.cell(row, 19).value): normalize(heloc.cell(row, 20).value)
                for row in range(21, 26)
                if heloc.cell(row, 19).value
            },
        },
    },
}

OUTPUT.write_text(json.dumps(data, indent=2) + '\n')
print(f'Wrote {OUTPUT}')
