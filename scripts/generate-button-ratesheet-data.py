from __future__ import annotations

import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE_ROOT = Path('/Users/ava/Documents/GitHub/first-access-lending')
WORKBOOK = SOURCE_ROOT / 'getaccess' / 'ratesheets' / 'latest_button.xlsx'
OUT = ROOT / 'src' / 'lib' / 'rates' / 'button-ratesheet.json'


def row_values(ws, row: int, cols=range(1, 14)):
    values = []
    for c in cols:
        v = ws.cell(row, c).value
        if v is not None:
            values.append(v)
    return values


def matrix(ws, row_start: int, row_end: int, col_start: int = 7, col_end: int = 13):
    return [
        [ws.cell(r, c).value for c in range(col_start, col_end + 1)]
        for r in range(row_start, row_end + 1)
    ]


def extract_labeled_matrix(ws, row_start: int, row_end: int, label_col: int = 6, value_col_start: int = 7, value_col_end: int = 13):
    rows = []
    values = []
    for r in range(row_start, row_end + 1):
        label = ws.cell(r, label_col).value
        if label is None:
            continue
        rows.append(label)
        values.append([ws.cell(r, c).value for c in range(value_col_start, value_col_end + 1)])
    return rows, values


def main() -> None:
    wb = load_workbook(WORKBOOK, data_only=False)
    ws = wb['Fully Delegated']

    note_rates = []
    for r in range(13, 53):
        rate = ws.cell(r, 1).value
        if rate is None:
            continue
        note_rates.append({
            'noteRate': rate,
            'prices': {
                'fullDoc': {
                    'HELOC': ws.cell(r, 2).value,
                    'CES': ws.cell(r, 3).value,
                },
                'altDoc': {
                    'HELOC': ws.cell(r, 4).value,
                    'CES': ws.cell(r, 5).value,
                },
            },
        })

    full_doc_fico_rows, full_doc_fico_values = extract_labeled_matrix(ws, 13, 21)
    alt_doc_fico_rows, alt_doc_fico_values = extract_labeled_matrix(ws, 26, 34)
    dti_rows, dti_values = extract_labeled_matrix(ws, 38, 40)
    balance_rows, balance_values = extract_labeled_matrix(ws, 52, 58)
    occupancy_rows, occupancy_values = extract_labeled_matrix(ws, 61, 63)
    bank_statement_rows, bank_statement_values = extract_labeled_matrix(ws, 64, 65)

    data = {
        'sourceWorkbook': str(WORKBOOK),
        'sheet': 'Fully Delegated',
        'guideMaxPrice': {
            'default': ws.cell(5, 7).value,
            'over500k': {
                'CES': ws.cell(5, 10).value,
                'HELOC': ws.cell(6, 10).value,
            },
        },
        'noteRates': note_rates,
        'tables': {
            'cltv': {
                'fullDocRows': full_doc_fico_rows,
                'altDocRows': alt_doc_fico_rows,
                'fullDocColumns': [ws.cell(12, c).value for c in range(7, 14)],
                'altDocColumns': [ws.cell(25, c).value for c in range(7, 14)],
                'fullDoc': matrix(ws, 13, 21),
                'altDoc': matrix(ws, 26, 34),
            },
            'fico': {
                'fullDoc': {
                    'rows': full_doc_fico_rows,
                    'columns': [ws.cell(12, c).value for c in range(7, 14)],
                    'values': full_doc_fico_values,
                },
                'altDoc': {
                    'rows': alt_doc_fico_rows,
                    'columns': [ws.cell(25, c).value for c in range(7, 14)],
                    'values': alt_doc_fico_values,
                },
            },
            'cashOut': {
                'rows': [ws.cell(r, 6).value for r in range(37, 40)],
                'columns': [ws.cell(36, c).value for c in range(7, 14)],
                'values': matrix(ws, 37, 39),
            },
            'dti': {
                'rows': dti_rows,
                'columns': [ws.cell(12, c).value for c in range(7, 14)],
                'values': dti_values,
            },
            'draw': {
                'heloc': {
                    'rows': [ws.cell(r, 6).value for r in range(44, 45)],
                    'columns': [ws.cell(12, c).value for c in range(7, 14)],
                    'values': matrix(ws, 44, 44),
                },
                'nonHeloc': {
                    'rows': [ws.cell(r, 6).value for r in range(45, 48)],
                    'columns': [ws.cell(12, c).value for c in range(7, 14)],
                    'values': matrix(ws, 45, 47),
                },
            },
            'maturity': {
                'rows': [ws.cell(r, 6).value for r in range(48, 52)],
                'columns': [ws.cell(12, c).value for c in range(7, 14)],
                'values': matrix(ws, 48, 51),
            },
            'balance': {
                'rows': balance_rows,
                'columns': [ws.cell(51, c).value for c in range(7, 14)],
                'values': balance_values,
            },
            'borrowerCount': {
                'rows': [ws.cell(r, 1).value for r in range(59, 61)],
                'columns': [ws.cell(59, c).value for c in range(7, 14)],
                'values': matrix(ws, 59, 60),
            },
            'selfEmployed': {
                'rows': [ws.cell(r, 1).value for r in range(59, 61)],
                'columns': [ws.cell(59, c).value for c in range(7, 14)],
                'values': matrix(ws, 59, 60),
            },
            'occupancy': {
                'rows': occupancy_rows,
                'columns': [ws.cell(60, c).value for c in range(7, 14)],
                'values': occupancy_values,
            },
            'unitCount': {
                'rows': occupancy_rows,
                'columns': [ws.cell(60, c).value for c in range(7, 14)],
                'values': occupancy_values,
            },
            'bankStatements': {
                'rows': bank_statement_rows,
                'columns': [ws.cell(12, c).value for c in range(7, 14)],
                'values': bank_statement_values,
            },
        },
    }

    OUT.write_text(json.dumps(data, indent=2), encoding='utf-8')
    print(f'Wrote {OUT}')


if __name__ == '__main__':
    main()
